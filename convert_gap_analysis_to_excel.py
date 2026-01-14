import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def convert_md_to_excel(md_file_path, excel_output_path):
    """
    Convert gap analysis markdown file to formatted Excel workbook
    """
    
    # Read the markdown file
    with open(md_file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Extract the table data
    table_data = []
    lines = content.split('\n')
    
    # Find the table start
    table_start = -1
    for i, line in enumerate(lines):
        if '| Category | spec.md UC Specification | models.md Implementation | Gap/Issue | Priority | Notes |' in line:
            table_start = i + 2  # Skip header and separator
            break
    
    if table_start == -1:
        print("Table not found in markdown file")
        return
    
    # Parse table rows
    in_table = False
    for i in range(table_start, len(lines)):
        line = lines[i].strip()
        
        # Skip empty lines and section headers
        if not line or line.startswith('##'):
            continue
            
        if line.startswith('|') and line.endswith('|'):
            # Split by | and clean up
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            
            # Skip separator row and ensure we have enough columns
            if len(cells) >= 6 and not cells[0].startswith('-'):
                # Clean up the data - remove markdown formatting
                category = re.sub(r'\*\*', '', cells[0]).strip()
                spec_req = re.sub(r'\*\*', '', cells[1]).strip()
                models_status = re.sub(r'✅|❌|\*\*', '', cells[2]).strip()
                gap_issue = re.sub(r'\*\*', '', cells[3]).strip()
                priority = re.sub(r'\*\*', '', cells[4]).strip()
                notes = re.sub(r'\*\*', '', cells[5]).strip()
                
                # Handle special status symbols
                if '✅' in cells[2]:
                    models_status = 'Implemented'
                elif '❌' in cells[2]:
                    models_status = 'Missing'
                
                # Skip summary rows
                if category and spec_req and not category.startswith('Summary'):
                    table_data.append([category, spec_req, models_status, gap_issue, priority, notes])
        
        # Stop at summary section
        if line.startswith('## Summary Statistics'):
            break
    
    # Create DataFrame
    df = pd.DataFrame(table_data, columns=['Category', 'spec.md UC Specification', 'models.md Implementation', 'Gap/Issue', 'Priority', 'Notes'])
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Analysis"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # Priority colors
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light Yellow
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green
    
    # Status colors
    implemented_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light Yellow
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = ['Category', 'spec.md UC Specification', 'models.md Implementation', 'Gap/Issue', 'Priority', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Apply priority colors
            if col_num == 5:  # Priority column
                if 'High' in str(value):
                    cell.fill = high_fill
                    cell.font = Font(bold=True)
                elif 'Medium' in str(value):
                    cell.fill = medium_fill
                elif 'Low' in str(value):
                    cell.fill = low_fill
            
            # Apply status colors
            elif col_num == 3:  # Status column
                if 'Implemented' in str(value):
                    cell.fill = implemented_fill
                elif 'Missing' in str(value):
                    cell.fill = missing_fill
                    cell.font = Font(bold=True)
                elif 'Partially' in str(value):
                    cell.fill = partial_fill
    
    # Adjust column widths
    column_widths = [15, 40, 20, 35, 12, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Add summary statistics section
    summary_start_row = len(table_data) + 4
    
    # Extract summary statistics
    summary_lines = []
    for line in lines:
        if '## Summary Statistics' in line:
            idx = lines.index(line)
            for j in range(idx + 1, len(lines)):
                if lines[j].startswith('##'):
                    break
                summary_lines.append(lines[j].strip())
            break
    
    # Add summary title
    summary_cell = ws.cell(row=summary_start_row, column=1, value="Summary Statistics")
    summary_cell.font = Font(bold=True, size=14)
    summary_cell.alignment = Alignment(horizontal='left')
    
    # Add summary data
    for i, line in enumerate(summary_lines, 1):
        if line.startswith('-'):
            clean_line = line.replace('- **', '').replace('**', '').replace('*', '')
            ws.cell(row=summary_start_row + i, column=1, value=clean_line)
    
    # Add critical gaps section
    critical_start_row = summary_start_row + len(summary_lines) + 3
    
    # Extract critical gaps
    critical_lines = []
    for line in lines:
        if '## Critical Priority Gaps (High Priority)' in line:
            idx = lines.index(line)
            for j in range(idx + 1, len(lines)):
                if lines[j].startswith('##'):
                    break
                if lines[j].strip().startswith(str(idx + 1) + '.'):
                    critical_lines.append(lines[j].strip())
            break
    
    # Add critical gaps title
    critical_cell = ws.cell(row=critical_start_row, column=1, value="Critical Priority Gaps")
    critical_cell.font = Font(bold=True, size=14, color="FF0000")
    critical_cell.alignment = Alignment(horizontal='left')
    
    # Add critical gaps data
    for i, line in enumerate(critical_lines, 1):
        ws.cell(row=critical_start_row + i, column=1, value=line)
    
    # Save the workbook
    wb.save(excel_output_path)
    print(f"Excel file saved successfully: {excel_output_path}")
    print(f"Total rows processed: {len(table_data)}")

if __name__ == "__main__":
    # File paths
    md_file = "gap_analysis_models_vs_spec_corrected.md"
    excel_file = "gap_analysis_models_vs_spec_final.xlsx"
    
    # Convert to Excel
    convert_md_to_excel(md_file, excel_file)
